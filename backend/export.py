from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def create_payments_xlsx(rows):
    """Создаёт Excel-файл из списка выплат."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Выплаты"

    headers = ["№", "Организация", "Сумма", "Дата", "Дата записи"]
    header_fill = PatternFill(start_color="16334B", end_color="16334B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for idx, row in enumerate(rows, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=row["organization"])
        ws.cell(row=idx + 1, column=3, value=row["amount"])
        ws.cell(row=idx + 1, column=4, value=row["date"])
        ws.cell(row=idx + 1, column=5, value=row["created_at"])

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20

    total = sum(r["amount"] for r in rows)
    last_row = len(rows) + 2
    ws.cell(row=last_row, column=2, value="ИТОГО:").font = Font(bold=True)
    ws.cell(row=last_row, column=3, value=total).font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
